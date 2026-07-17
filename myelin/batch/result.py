from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING, Any

from pydantic import BaseModel, ConfigDict, Field

if TYPE_CHECKING:
    from myelin.batch.options import BatchOptions
    from myelin.core import MyelinIO, MyelinOutput
else:
    from myelin.batch.options import BatchOptions  # noqa: F401  (needed for Pydantic forward-ref resolution)


PRICER_OUTPUT_ATTRS: tuple[tuple[str, str], ...] = (
    ("ipps", "total_payment"),
    ("psych", "total_payment"),
    ("ltch", "total_payment"),
    ("irf", "total_payment"),
    ("hospice", "total_payment"),
    ("snf", "total_payment"),
    ("hha", "total_payment"),
    ("esrd", "total_payment"),
    ("fqhc", "total_payment"),
    ("asc", "total_payment"),
    ("opps", "total_claim_payment"),
)

PER_CLAIM_DETAIL_SHEET_LIMIT = 50


def extract_total_payment(output: "MyelinOutput | None") -> float:
    if output is None:
        return 0.0
    total = 0.0
    for attr, field_name in PRICER_OUTPUT_ATTRS:
        pricer_output = getattr(output, attr, None)
        if pricer_output is None:
            continue
        value = getattr(pricer_output, field_name, None)
        if value is None:
            continue
        try:
            total += float(value)
        except (TypeError, ValueError):
            continue
    return total


def extract_per_pricer_totals(output: "MyelinOutput | None") -> dict[str, float]:
    if output is None:
        return {}
    totals: dict[str, float] = {}
    for attr, field_name in PRICER_OUTPUT_ATTRS:
        pricer_output = getattr(output, attr, None)
        if pricer_output is None:
            continue
        value = getattr(pricer_output, field_name, None)
        if value is None:
            continue
        try:
            totals[attr] = float(value)
        except (TypeError, ValueError):
            continue
    return totals


def classify_error(output: "MyelinOutput | None") -> str | None:
    if output is None:
        return "MyelinOutput is None"
    if output.error:
        return str(output.error)
    return None


class BatchStats(BaseModel):
    """Aggregate statistics for a completed batch run."""

    model_config = ConfigDict(extra="forbid")

    total_count: int = 0
    success_count: int = 0
    failure_count: int = 0
    skipped_count: int = 0
    elapsed_seconds: float = 0.0
    claims_per_second: float = 0.0
    total_payment: float = 0.0
    per_pricer_total_payment: dict[str, float] = Field(default_factory=dict)
    error_histogram: dict[str, int] = Field(default_factory=dict)
    started_at: datetime | None = None
    finished_at: datetime | None = None


def _mio_to_dict(mio: "MyelinIO") -> dict[str, Any]:
    record: dict[str, Any] = {}
    if mio.input is not None:
        record["input"] = mio.input.model_dump(mode="json")
    if mio.output is not None:
        record["output"] = mio.output.model_dump(mode="json")
    return record


def _safe_sheet_name(index: int) -> str:
    return f"Claim_{index + 1:04d}"[:31]


def _build_workbook(result: "BatchResult"):
    from openpyxl import Workbook

    from myelin.helpers.excel_exporter import (
        _ensure_openpyxl,
        _flatten_model,
        _format_value,
    )

    _ensure_openpyxl()

    wb = Workbook()
    summary_ws = wb.active
    assert summary_ws is not None
    summary_ws.title = "BatchSummary"
    summary_ws.append(["Metric", "Value"])
    summary_ws.append(["total_count", result.stats.total_count])
    summary_ws.append(["success_count", result.stats.success_count])
    summary_ws.append(["failure_count", result.stats.failure_count])
    summary_ws.append(["skipped_count", result.stats.skipped_count])
    summary_ws.append(["elapsed_seconds", result.stats.elapsed_seconds])
    summary_ws.append(["claims_per_second", result.stats.claims_per_second])
    summary_ws.append(["total_payment", result.stats.total_payment])
    for attr, value in result.stats.per_pricer_total_payment.items():
        summary_ws.append([f"payment_{attr}", value])
    for err, count in result.stats.error_histogram.items():
        summary_ws.append([f"error:{err}", count])

    per_claim_ws = wb.create_sheet("PerClaim")
    per_claim_ws.append(
        [
            "index",
            "claim_id",
            "status",
            "error",
            "total_payment",
        ]
    )
    for i, mio in enumerate(result.items):
        error = classify_error(mio.output)
        claim_id = mio.input.claimid if mio.input is not None else ""
        per_claim_ws.append(
            [
                i,
                claim_id,
                "error" if error else "ok",
                error or "",
                extract_total_payment(mio.output),
            ]
        )

    for i, mio in enumerate(result.items):
        if i >= PER_CLAIM_DETAIL_SHEET_LIMIT:
            break
        if mio.output is None:
            continue
        ws = wb.create_sheet(_safe_sheet_name(i))
        flat = _flatten_model(mio.output)
        ws.append(["field", "value"])
        for key, value in flat.items():
            ws.append([key, _format_value(value)])

    return wb


class BatchResult(BaseModel):
    """Container for the results of Myelin.process_batch()."""

    model_config = ConfigDict(extra="forbid")

    items: list["MyelinIO"] = Field(default_factory=list)
    stats: BatchStats = Field(default_factory=BatchStats)
    options: "BatchOptions | None" = None

    def succeeded(self) -> list["MyelinIO"]:
        return [mio for mio in self.items if classify_error(mio.output) is None]

    def failed(self) -> list["MyelinIO"]:
        return [mio for mio in self.items if classify_error(mio.output) is not None]

    def to_jsonl(self, path: str | Path) -> None:
        output_path = Path(path)
        with output_path.open("w", encoding="utf-8") as f:
            for mio in self.items:
                f.write(json.dumps(_mio_to_dict(mio), default=str) + "\n")

    def to_jsonl_bytes(self) -> bytes:
        if not self.items:
            return b""
        return (
            "\n".join(
                json.dumps(_mio_to_dict(mio), default=str) for mio in self.items
            )
            + "\n"
        ).encode("utf-8")

    def to_excel(self, path: str | Path) -> None:
        wb = _build_workbook(self)
        wb.save(str(path))

    def to_excel_bytes(self) -> bytes:
        wb = _build_workbook(self)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()


__all__ = [
    "BatchStats",
    "BatchResult",
    "extract_total_payment",
    "extract_per_pricer_totals",
    "classify_error",
]
